import openpyxl
from openpyxl.styles import Alignment, Font
import pandas as pd

# File paths
ITEM_BATCH_REPORT_PATH = r'D:\OneDrive\Desktop\VYAPAR\Item Batch Report.xlsx'
EXPORT_ITEMS_PATH = r'D:\OneDrive\Desktop\VYAPAR\Export Items.xlsx'
HTML_OUTPUT_PATH = r'D:\OneDrive\Desktop\VYAPAR\Item_Batch_Report.html'

def update_excel_file():
    """Updates the Excel file by adding columns and mapping data."""
    # Load workbooks and worksheets
    item_batch_report_wb = openpyxl.load_workbook(ITEM_BATCH_REPORT_PATH)
    export_items_wb = openpyxl.load_workbook(EXPORT_ITEMS_PATH)
    
    item_batch_report_ws = item_batch_report_wb.active  # Assuming the active sheet is correct
    export_items_ws = export_items_wb.active  # Assuming the active sheet is correct

    # Step 1: Delete the 2nd row in 'Item Batch Report.xlsx'
    item_batch_report_ws.delete_rows(2, 1)  # Deletes the 2nd row

    # Step 2: Add new columns (H to L) with specified headers
    headers = ["Company", "Description", "Category", "Tax Rate", "Location"]
    for col_num, header in enumerate(headers, start=8):  # Starting from Column H (index 8)
        cell = item_batch_report_ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, size=12)  # Bold and larger font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Step 3: Extract data from 'Export Items.xlsx' based on unique values in Column A
    export_items_data = {
        row[0]: {
            "Company": row[5],         # Column F from Export Items.xlsx
            "Description": row[2],    # Column C
            "Category": row[3],       # Column D
            "Tax Rate": row[14],      # Column O
            "Location": row[13]       # Column N
        }
        for row in export_items_ws.iter_rows(min_row=2, values_only=True)  # Data starts from row 2
        if row[0]  # Ensures 'Item Name' (Column A) is not None
    }

    # Step 4: Populate data in the Item Batch Report
    for row in item_batch_report_ws.iter_rows(min_row=2, max_row=item_batch_report_ws.max_row, values_only=False):
        item_name = row[0].value  # Column A value
        if item_name in export_items_data:
            # Map data from 'Export Items.xlsx' to Columns H to L
            export_data = export_items_data[item_name]
            row[7].value = export_data["Company"]      # Column H
            row[8].value = export_data["Description"] # Column I
            row[9].value = export_data["Category"]    # Column J
            row[10].value = export_data["Tax Rate"]   # Column K
            row[11].value = export_data["Location"]   # Column L

    # Save the updated workbook
    item_batch_report_wb.save(ITEM_BATCH_REPORT_PATH)
    print(f"Updated Excel file saved successfully at '{ITEM_BATCH_REPORT_PATH}'.")

def generate_html_view():
    """Generates an HTML view with styled headers and added columns."""
    # Load data into DataFrame
    df = pd.read_excel(ITEM_BATCH_REPORT_PATH, sheet_name=0).dropna(how="all")  # Load data from first sheet
    
    # Strip unnecessary whitespace from headers
    df.columns = df.columns.str.strip()
    
    # Apply formatting (e.g., numeric columns with 2 decimal places)
    if 'MRP' in df.columns:
        df['MRP'] = df['MRP'].apply(lambda x: f"{x:.2f}" if isinstance(x, (int, float)) else x)
    if 'Purchase price' in df.columns:
        df['Purchase price'] = df['Purchase price'].apply(lambda x: f"{x:.2f}" if isinstance(x, (int, float)) else x)
    
    # Generate HTML with uniform column widths and proper data inclusion
    html_template = f"""
    <!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Item Batch Report</title>
        <script src="https://code.jquery.com/jquery-3.6.0.min.js"></script>
        <script src="https://cdn.datatables.net/1.13.6/js/jquery.dataTables.min.js"></script>
        <script src="https://cdn.datatables.net/fixedheader/3.3.2/js/dataTables.fixedHeader.min.js"></script>
        <link rel="stylesheet" href="https://cdn.datatables.net/1.13.6/css/jquery.dataTables.min.css">
        <style>
            body {{
                font-family: Arial, sans-serif;
                margin: 20px;
            }}
            table {{
                width: 100%;
                margin: auto;
                border-collapse: collapse;
            }}
            th {{
                background-color: #4CAF50; /* Green header background */
                color: white; /* White text for headers */
                font-weight: bold;
                text-align: center;
                padding: 8px;
            }}
            td {{
                text-align: center; /* Center-align data */
                vertical-align: middle;
                padding: 8px;
                border: 1px solid #ddd;
                word-wrap: break-word; /* Enable text wrapping for long text */
            }}
            tr:nth-child(even) {{
                background-color: #f9f9f9; /* Light gray for even rows */
            }}
            tr:nth-child(odd) {{
                background-color: #ffffff; /* White for odd rows */
            }}
            table.dataTable th,
            table.dataTable td {{
                width: 10%; /* Set uniform width for all columns */
            }}
            .dataTables_wrapper .dataTables_filter {{
                float: right;
                text-align: right;
            }}
            .dataTables_wrapper .dataTables_length {{
                float: left;
            }}
            table.dataTable thead th {{
                position: sticky;
                top: 0;
                z-index: 100;
            }}
        </style>
    </head>
    <body>
        <h2>Item Batch Report</h2>
        <table id="itemBatchTable" class="display">
            <thead>
                <tr>
                    {"".join(f"<th>{col}</th>" for col in df.columns)} <!-- Include all column headers -->
                </tr>
            </thead>
            <tbody>
                {"".join(f"<tr>{''.join(f'<td>{val if pd.notna(val) else ""}</td>' for val in row)}</tr>" for row in df.values)}
            </tbody>
        </table>
        <script>
            $(document).ready(function() {{
                $('#itemBatchTable').DataTable({{
                    fixedHeader: true,
                    dom: 'lfrtip',  // Length menu, filter, table, pagination, info
                    pageLength: 10,  // Display 10 records per page
                    order: [],
                }});
            }});
        </script>
    </body>
    </html>
    """
    # Save the HTML file
    with open(HTML_OUTPUT_PATH, "w", encoding="utf-8") as file:
        file.write(html_template)
    print(f"HTML view created successfully at '{HTML_OUTPUT_PATH}'.")

if __name__ == "__main__":
    update_excel_file()
    generate_html_view()